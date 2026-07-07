import os
import json
from typing import List, Dict
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from SE_to_PLM.core.models.export_row import ExportRow
from SE_to_PLM.core.services.export.excel_styles import apply_header_style, apply_data_style
from SE_to_PLM.core.services.export.column_auto_size import auto_size_columns
from SE_to_PLM.core.services.plm.abbreviation_service import abbreviation_service
from SE_to_PLM.core.services.logging.logger_service import logger

class ExcelExporter:
    """
    Service for generating dynamically configured Excel files from ExportRow data.
    """

    def load_columns_config(self) -> List[dict]:
        """Loads column configuration from JSON file or returns default."""
        config_path = Path(__file__).parent.parent.parent.parent / "ui" / "resources" / "columns_config.json"
        if config_path.exists():
            try:
                with open(config_path, "r", encoding="utf-8") as f:
                    return json.load(f)
            except Exception as e:
                logger.error(f"Erreur lors du chargement de columns_config.json : {e}")
        return self.get_default_columns_config()

    def get_default_columns_config(self) -> List[dict]:
        """Returns the default column configuration matching original headers."""
        return [
            {"header": "Level", "source_type": "special_processed", "source_name": "level", "default_value": "0", "style": "cad"},
            {"header": "Relationship", "source_type": "special_processed", "source_name": "relationship", "default_value": "", "style": "cad"},
            {"header": "ordre", "source_type": "special_processed", "source_name": "order", "default_value": "", "style": "cad"},
            {"header": "quantite", "source_type": "special_processed", "source_name": "quantity", "default_value": "1", "style": "cad"},
            {"header": "repere", "source_type": "special_processed", "source_name": "repere", "default_value": "", "style": "cad"},
            {"header": "SpecialCAD", "source_type": "special_processed", "source_name": "special_cad", "default_value": "", "style": "cad"},
            {"header": "Class", "source_type": "special_processed", "source_name": "plm_class", "default_value": "", "style": "cad"},
            {"header": "ref_utilisat", "source_type": "special_processed", "source_name": "ref_utilisat", "default_value": "", "style": "cad"},
            {"header": "version", "source_type": "special_processed", "source_name": "version", "default_value": "-", "style": "cad"},
            {"header": "indice_1", "source_type": "special_processed", "source_name": "indice_1", "default_value": "-", "style": "cad"},
            {"header": "indice_2", "source_type": "special_processed", "source_name": "indice_2", "default_value": "-", "style": "cad"},
            {"header": "revision", "source_type": "special_processed", "source_name": "revision", "default_value": "1", "style": "cad"},
            {"header": "designation", "source_type": "special_processed", "source_name": "designation", "default_value": "", "style": "plm"},
            {"header": "designation_erp", "source_type": "special_processed", "source_name": "designation_erp", "default_value": "", "style": "plm"},
            {"header": "cus_createur", "source_type": "solid_edge_property", "source_name": "auteur", "default_value": "", "style": "plm"},
            {"header": "cus_date_crea", "source_type": "solid_edge_property", "source_name": "date_creation", "default_value": "", "style": "plm"},
            {"header": "user_version_1", "source_type": "solid_edge_property", "source_name": "auteur_modif", "default_value": "", "style": "plm"},
            {"header": "date_version_1", "source_type": "solid_edge_property", "source_name": "date_modif", "default_value": "", "style": "plm"},
            {"header": "matiere", "source_type": "solid_edge_property", "source_name": "matiere", "default_value": "", "style": "plm"},
            {"header": "densite", "source_type": "solid_edge_property", "source_name": "densite", "default_value": "", "style": "plm"},
            {"header": "dia_se", "source_type": "solid_edge_property", "source_name": "dia_se", "default_value": "", "style": "plm"},
            {"header": "mode_appro", "source_type": "special_processed", "source_name": "mode_appro", "default_value": "", "style": "plm"},
            {"header": "Attachments", "source_type": "special_processed", "source_name": "attachments", "default_value": "", "style": "plm"}
        ]

    def create_export(self, rows: List[ExportRow], output_path: str,
                      optimize_designations: bool = False,
                      highlight_modifications: bool = False):
        """
        Creates a new Excel workbook, writes the data based on columns_config.json,
        applies styles, and saves it.
        """
        logger.info(f"Création du fichier Excel...")
        
        # Ensure directory exists
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Structure"

        # Load columns config
        columns_config = self.load_columns_config()

        # 1. Write Headers
        for col_idx, col in enumerate(columns_config, 1):
            cell = ws.cell(row=1, column=col_idx, value=col["header"])
            # Apply dynamic colors (Green for CAD, Orange for PLM)
            from SE_to_PLM.core.services.export.excel_styles import STYLE_HEADER_CAD, STYLE_HEADER_PLM
            style = STYLE_HEADER_PLM if col.get("style") == "plm" else STYLE_HEADER_CAD
            cell.font = style["font"]
            cell.fill = style["fill"]
            cell.alignment = style["alignment"]

        # 2. Write Data Rows
        fill_modified = PatternFill(fill_type="solid", fgColor="FF9999")
        
        for row_idx, row_obj in enumerate(rows, 2):
            row_was_modified = False
            
            # Extract designation value
            orig_designation = row_obj.special_values.get("designation", "")
            
            optimized_erp = orig_designation
            if optimize_designations and isinstance(orig_designation, str):
                optimized, modified = abbreviation_service.optimiser_designation(orig_designation)
                optimized_erp = abbreviation_service.majuscules_sans_accents(optimized)
                if optimized_erp != orig_designation:
                    row_was_modified = True
                    
            for col_idx, col in enumerate(columns_config, 1):
                if col["source_type"] == "special_processed" and col["source_name"] == "designation_erp":
                    value = optimized_erp
                else:
                    value = row_obj.get_value(col)
                    
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                apply_data_style(cell)
                
            if row_was_modified and highlight_modifications:
                for col_idx in range(1, len(columns_config) + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = fill_modified

        # 3. Auto-size columns
        try:
            auto_size_columns(ws)
        except Exception:
            pass

        # 4. Save
        try:
            # Ensure .xlsx extension
            if not output_path.lower().endswith(".xlsx"):
                output_path += ".xlsx"
                
            wb.save(output_path)
        except Exception as e:
            logger.error(f"Erreur d'enregistrement Excel : {e}")
            raise

# Global instance
excel_exporter = ExcelExporter()
