from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side

# Colors
COLOR_HEADER_CAD = "CCFFCC" # Light Green
COLOR_HEADER_PLM = "FFC000" # Orange

# Styles
STYLE_HEADER_CAD = {
    "font": Font(bold=True, name="Segoe UI", size=10),
    "fill": PatternFill(start_color=COLOR_HEADER_CAD, end_color=COLOR_HEADER_CAD, fill_type="solid"),
    "alignment": Alignment(horizontal="left", vertical="center")
}

STYLE_HEADER_PLM = {
    "font": Font(bold=True, name="Segoe UI", size=10),
    "fill": PatternFill(start_color=COLOR_HEADER_PLM, end_color=COLOR_HEADER_PLM, fill_type="solid"),
    "alignment": Alignment(horizontal="left", vertical="center")
}

STYLE_DATA = {
    "font": Font(name="Segoe UI", size=10),
    "alignment": Alignment(horizontal="left", vertical="center")
}

def apply_header_style(cell, column_index: int):
    """
    Applies the appropriate header style based on the column index (1-based).
    Columns 1-12: CAD Style (Green)
    Columns 13-21: PLM Style (Orange)
    """
    style = STYLE_HEADER_CAD if column_index <= 12 else STYLE_HEADER_PLM
    cell.font = style["font"]
    cell.fill = style["fill"]
    cell.alignment = style["alignment"]

def apply_data_style(cell):
    """Applies the standard data style to a cell."""
    cell.font = STYLE_DATA["font"]
    cell.alignment = STYLE_DATA["alignment"]
