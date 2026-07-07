import json
import os
import re
import unicodedata
from pathlib import Path
from typing import List, Dict, Tuple, Optional
from SE_to_PLM.core.services.logging.logger_service import logger

class AbbreviationService:
    """
    Service gérant le dictionnaire d'abréviations (chargement, sauvegarde, modifications)
    et appliquant les règles d'abréviation et de limitation à 32 caractères.
    """
    
    LIMITE_CARACTERES = 32
    MOTS_A_PRESERVER = ["VINOLOK", "EMF", "GALAXY", "NEO"]

    def __init__(self):
        self.resources_dir = Path(__file__).parent.parent.parent.parent / "ui" / "resources"
        self.json_path = self.resources_dir / "abbreviations.json"
        self.abbreviations: List[dict] = []
        self.compiled_rules = []
        self.load_abbreviations()

    def get_default_abbreviations(self) -> List[dict]:
        """Retourne la liste d'abréviations par défaut."""
        # On peut importer temporairement le dictionnaire brut si disponible pour ne pas dupliquer
        # Mais pour être indépendant et résilient, on définit une liste de secours au cas où.
        # En pratique, le fichier abbreviations.json a déjà été généré par bootstrap,
        # mais on gère le cas vide.
        return []

    def load_abbreviations(self):
        """Charge les abréviations depuis le fichier JSON."""
        if not self.json_path.exists():
            # Initialisation par défaut vide, le bootstrap s'occupe de le peupler.
            self.abbreviations = []
            self.compiled_rules = []
            return

        try:
            with open(self.json_path, "r", encoding="utf-8") as f:
                self.abbreviations = json.load(f)
            self._compile_rules()
            logger.info(f"Dictionnaire d'abréviations chargé : {len(self.abbreviations)} entrées.")
        except Exception as e:
            logger.error(f"Erreur lors du chargement des abréviations : {e}")
            self.abbreviations = []
            self.compiled_rules = []

    def save_abbreviations(self, abbreviations: List[dict]):
        """Enregistre les abréviations dans le fichier JSON et recharge les règles."""
        try:
            # S'assurer que le dossier existe
            self.resources_dir.mkdir(parents=True, exist_ok=True)
            
            with open(self.json_path, "w", encoding="utf-8") as f:
                json.dump(abbreviations, f, ensure_ascii=False, indent=2)
            
            self.abbreviations = abbreviations
            self._compile_rules()
            logger.info(f"Dictionnaire d'abréviations sauvegardé avec succès ({len(self.abbreviations)} entrées).")
        except Exception as e:
            logger.error(f"Erreur lors de la sauvegarde des abréviations : {e}")
            raise

    def _supprimer_accents(self, texte: str) -> str:
        texte = unicodedata.normalize("NFD", str(texte))
        return "".join(
            caractere
            for caractere in texte
            if unicodedata.category(caractere) != "Mn"
        )

    def _variantes_terme(self, terme: str) -> List[str]:
        """Génère les variantes d'un terme (accents, pluriels s/x)."""
        terme = re.sub(r"\s+", " ", str(terme)).strip()
        variantes = []

        morceaux = [m.strip() for m in terme.split("/") if m.strip()] if " / " in terme else [terme]

        for morceau in morceaux:
            candidats = [morceau]

            if "(x)" in morceau:
                candidats.append(morceau.replace("(x)", ""))
                candidats.append(morceau.replace("(x)", "x"))
            if "(s)" in morceau:
                candidats.append(morceau.replace("(s)", ""))
                candidats.append(morceau.replace("(s)", "s"))

            for candidat in candidats:
                candidat = candidat.replace("(x)", "").replace("(s)", "")
                candidat = re.sub(r"\s+", " ", candidat).strip()
                if not candidat:
                    continue
                variantes.append(candidat)
                variantes.append(self._supprimer_accents(candidat))

        # Dédoublonnage ordonné
        resultat = []
        vus = set()
        for variante in variantes:
            cle = variante.casefold()
            if cle not in vus:
                vus.add(cle)
                resultat.append(variante)
        return resultat

    def _compile_rules(self):
        """Compile et trie les expressions régulières pour application rapide."""
        temp_rules = []
        vus = set()

        for item in self.abbreviations:
            terme = item.get("terme", "").strip()
            abbrev = item.get("abreviation", "").strip()
            priorite = item.get("priorite", 2)
            
            if not terme or not abbrev:
                continue

            for variante in self._variantes_terme(terme):
                cle = (variante.casefold(), abbrev.casefold(), priorite)
                if cle in vus:
                    continue
                vus.add(cle)
                temp_rules.append((variante, abbrev, priorite))

        # Tri : priorité croissante, puis gain de caractères décroissant
        self.compiled_rules = []
        for texte_long, texte_court, priorite in sorted(
            temp_rules,
            key=lambda x: (x[2], -(len(x[0]) - len(x[1]))),
        ):
            self.compiled_rules.append({
                "motif": re.compile(rf"\b{re.escape(texte_long)}\b", flags=re.IGNORECASE),
                "remplacement": texte_court,
                "texte_long": texte_long,
                "priorite": priorite,
                "gain": len(texte_long) - len(texte_court),
            })

    def majuscules_sans_accents(self, texte: str) -> str:
        """Convertit le texte en MAJUSCULES sans accents."""
        return self._supprimer_accents(texte).upper()

    def appliquer_abreviations_prioritaires(self, texte: str) -> Tuple[str, bool]:
        """Applique toutes les abréviations du dictionnaire de manière ordonnée (priorité puis gain)."""
        texte = re.sub(r"\s+", " ", str(texte)).strip()
        texte_modifie = texte
        modifie = False

        for regle in self.compiled_rules:
            nouveau_texte = regle["motif"].sub(
                regle["remplacement"],
                texte_modifie
            )
            if nouveau_texte != texte_modifie:
                texte_modifie = re.sub(r"\s+", " ", nouveau_texte).strip()
                modifie = True

        return texte_modifie, modifie

    def tronquer_mots_entiers(self, texte: str, limite: int = LIMITE_CARACTERES) -> str:
        """Tronque à la limite sans couper les mots."""
        texte = re.sub(r"\s+", " ", str(texte)).strip()

        if len(texte) <= limite:
            return texte

        mots = texte.split(" ")
        resultat = ""

        for mot in mots:
            candidat = mot if resultat == "" else f"{resultat} {mot}"

            if len(candidat) <= limite:
                resultat = candidat
            else:
                break

        return resultat

    def trouver_mot_a_preserver(self, texte: str) -> Optional[str]:
        """Retourne le premier mot-clé à préserver présent dans le texte."""
        for mot in self.MOTS_A_PRESERVER:
            if re.search(rf"\b{re.escape(mot)}\b", texte, flags=re.IGNORECASE):
                return mot
        return None

    def tronquer_en_conservant_mot_cle(self, texte: str, mot_cle: str, limite: int = LIMITE_CARACTERES) -> str:
        """Tronque en conservant obligatoirement le mot-clé et en gardant le maximum de mots avant."""
        texte = re.sub(r"\s+", " ", str(texte)).strip()

        match = re.search(
            rf"\b{re.escape(mot_cle)}\b",
            texte,
            flags=re.IGNORECASE,
        )

        if not match:
            return self.tronquer_mots_entiers(texte, limite)

        resultat = match.group(0)
        avant = texte[:match.start()].strip()
        mots_avant = avant.split() if avant else []

        for mot in reversed(mots_avant):
            candidat = f"{mot} {resultat}"

            if len(candidat) <= limite:
                resultat = candidat
            else:
                break

        return resultat

    def optimiser_designation(self, texte: str) -> Tuple[str, bool]:
        """
        Optimise une désignation : applique toutes les abréviations, puis
        tronque à 32 caractères si nécessaire.
        """
        if not texte:
            return "", False

        texte = re.sub(r"\s+", " ", str(texte)).strip()

        if len(texte) <= self.LIMITE_CARACTERES:
            return texte, False

        texte_abrege, abrege = self.appliquer_abreviations_prioritaires(texte)

        if len(texte_abrege) <= self.LIMITE_CARACTERES:
            return texte_abrege, abrege

        mot_cle = self.trouver_mot_a_preserver(texte_abrege)
        if mot_cle:
            return self.tronquer_en_conservant_mot_cle(texte_abrege, mot_cle), True

        return self.tronquer_mots_entiers(texte_abrege), True

    # ==========================================================
    # IMPORT / EXPORT EXCEL & ODS
    # ==========================================================

    def export_to_excel(self, file_path: str):
        """Exporte les abréviations actuelles au format Excel (.xlsx)."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Dictionnaire"
        
        headers = ["Terme", "Abréviation", "Priorité"]
        ws.append(headers)
        
        # Styles
        font_header = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        fill_header = PatternFill(fill_type="solid", fgColor="1F497D")
        alignment = Alignment(horizontal="center", vertical="center")
        
        for col_idx in range(1, 4):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = alignment
            
        # Data
        for item in sorted(self.abbreviations, key=lambda x: x.get("terme", "").lower()):
            ws.append([item.get("terme", ""), item.get("abreviation", ""), item.get("priorite", 2)])
            
        # Auto-size columns
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = col[0].column_letter
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
            
        wb.save(file_path)

    def import_from_excel_or_ods(self, file_path: str, mode: str = "merge") -> int:
        """
        Importe les abréviations depuis un fichier .xlsx ou .ods.
        mode : "merge" pour fusionner et dédoublonner, "replace" pour écraser complètement.
        Retourne le nombre d'entrées importées.
        """
        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"Fichier introuvable : {file_path}")
            
        suffix = path.suffix.lower()
        imported = []
        
        if suffix == ".xlsx":
            imported = self._read_xlsx_file(file_path)
        elif suffix == ".ods":
            imported = self._read_ods_file(file_path)
        else:
            raise ValueError("Format de fichier non supporté. Utilisez .xlsx ou .ods.")
            
        if not imported:
            return 0
            
        if mode == "replace":
            self.save_abbreviations(imported)
        else:  # merge
            # Créer un dictionnaire indexé par le terme normalisé
            existing_dict = {item["terme"].strip().lower(): item for item in self.abbreviations}
            
            for item in imported:
                key = item["terme"].strip().lower()
                # Les importations remplacent l'existant si elles existent, ou s'ajoutent
                existing_dict[key] = item
                
            new_list = list(existing_dict.values())
            self.save_abbreviations(new_list)
            
        return len(imported)

    def _read_xlsx_file(self, file_path: str) -> List[dict]:
        from openpyxl import load_workbook
        wb = load_workbook(file_path, data_only=True)
        imported = []
        
        for name in wb.sheetnames:
            ws = wb[name]
            # On vérifie la ligne d'en-tête pour savoir dans quelles colonnes se trouvent les infos
            rows_iter = ws.iter_rows(max_row=1)
            try:
                first_row = next(rows_iter)
                header = [str(cell.value or '').strip().lower() for cell in first_row]
            except StopIteration:
                continue
            
            col_terme = 0
            col_abbrev = 1
            col_priorite = 2
            
            # Tenter de trouver les colonnes par nom
            for idx, h in enumerate(header):
                if "terme" in h or "term" in h or "mot" in h:
                    col_terme = idx
                elif "abré" in h or "abre" in h or "short" in h:
                    col_abbrev = idx
                elif "prior" in h or "prio" in h:
                    col_priorite = idx

            # Lecture à partir de la ligne 2
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or len(row) <= max(col_terme, col_abbrev):
                    continue
                    
                terme = str(row[col_terme]).strip() if row[col_terme] is not None else ""
                abbrev = str(row[col_abbrev]).strip() if row[col_abbrev] is not None else ""
                
                if terme and abbrev:
                    priority = 2
                    if len(row) > col_priorite and row[col_priorite] is not None:
                        try:
                            priority = int(row[col_priorite])
                        except:
                            pass
                            
                    imported.append({
                        "terme": terme,
                        "abreviation": abbrev,
                        "priorite": priority
                    })
        return imported

    def _read_ods_file(self, file_path: str) -> List[dict]:
        import zipfile
        import xml.etree.ElementTree as ET
        
        with zipfile.ZipFile(file_path) as z:
            content_xml = z.read("content.xml")
            
        root = ET.fromstring(content_xml)
        namespaces = {
            'table': 'urn:oasis:names:tc:opendocument:xmlns:table:1.0',
            'text': 'urn:oasis:names:tc:opendocument:xmlns:text:1.0'
        }
        
        imported = []
        sheets = root.findall('.//table:table', namespaces)
        
        for sheet in sheets:
            rows = sheet.findall('.//table:table-row', namespaces)
            if not rows:
                continue
                
            # Lire l'en-tête (première ligne non vide)
            col_terme = 0
            col_abbrev = 1
            col_priorite = 2
            
            header_found = False
            
            for row in rows:
                cells = row.findall('.//table:table-cell', namespaces)
                row_values = []
                for cell in cells:
                    text_node = cell.find('.//text:p', namespaces)
                    val = text_node.text if text_node is not None else ""
                    row_values.append(val)
                while row_values and row_values[-1] == "":
                    row_values.pop()
                    
                if not row_values:
                    continue
                    
                if not header_found:
                    # Analyser l'en-tête
                    header_found = True
                    for idx, val in enumerate(row_values):
                        h = val.strip().lower()
                        if "terme" in h or "term" in h or "mot" in h:
                            col_terme = idx
                        elif "abré" in h or "abre" in h or "short" in h:
                            col_abbrev = idx
                        elif "prior" in h or "prio" in h:
                            col_priorite = idx
                    continue
                
                # Traitement des lignes de données
                if len(row_values) <= max(col_terme, col_abbrev):
                    continue
                    
                terme = str(row_values[col_terme]).strip()
                abbrev = str(row_values[col_abbrev]).strip()
                
                if terme and abbrev:
                    priority = 2
                    if len(row_values) > col_priorite and row_values[col_priorite]:
                        try:
                            priority = int(str(row_values[col_priorite]).strip())
                        except:
                            pass
                            
                    imported.append({
                        "terme": terme,
                        "abreviation": abbrev,
                        "priorite": priority
                    })
                    
        return imported

# Instance globale
abbreviation_service = AbbreviationService()
