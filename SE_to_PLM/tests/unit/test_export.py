import os
import pytest
from SE_to_PLM.core.models.export_row import ExportRow
from SE_to_PLM.core.services.export.excel_exporter import ExcelExporter
from openpyxl import load_workbook

def test_excel_generation(tmp_path):
    output_file = str(tmp_path / "test_export.xlsx")
    
    rows = [
        ExportRow(
            level=0, relationship="", order=1, quantity=1, repere="",
            special_cad="Root", plm_class="SUB_ASSY_A", ref_utilisat="Root",
            version="A", indice_1="-", indice_2="-", revision="1",
            designation="Top Assy", cus_createur="Admin", cus_date_crea="2026",
            user_version_1="", date_version_1="", matiere="",
            densite="", dia_se="", mode_appro="Fabrication interne", attachments="C:/Root.asm"
        ),
        ExportRow(
            level=1, relationship="ComposedOf", order=2, quantity=2, repere="",
            special_cad="Part1", plm_class="PART_A", ref_utilisat="Part1",
            version="B", indice_1="A", indice_2="-", revision="1",
            designation="Simple Part", cus_createur="User", cus_date_crea="2026",
            user_version_1="", date_version_1="", matiere="Steel",
            densite="7.8", dia_se="", mode_appro="Fabrication interne", attachments="C:/Part1.par"
        )
    ]
    
    exporter = ExcelExporter()
    exporter.create_export(rows, output_file)
    
    assert os.path.exists(output_file)
    
    wb = load_workbook(output_file)
    ws = wb.active
    assert ws.title == "Structure"
    assert ws.cell(row=1, column=1).value == "Level"
    assert ws.cell(row=1, column=13).value == "designation"
    assert ws.cell(row=1, column=14).value == "designation_erp"
    
    assert ws.cell(row=2, column=6).value == "Root"
    assert ws.cell(row=3, column=6).value == "Part1"
    assert ws.cell(row=3, column=4).value == 2 # quantity
    
    # Without optimization: both should have the original designation
    assert ws.cell(row=2, column=13).value == "Top Assy"
    assert ws.cell(row=2, column=14).value == "Top Assy"
    
    # Test with optimization
    output_file_opt = str(tmp_path / "test_export_opt.xlsx")
    
    # Change designation to be long so that it triggers abbreviation and truncation
    rows[1].designation = "Simple Part with a very long designation name that exceeds 32 characters"
    
    # Let's mock abbreviation rules for testing
    from SE_to_PLM.core.services.plm.abbreviation_service import abbreviation_service
    original_abbrevs = list(abbreviation_service.abbreviations)
    
    try:
        abbreviation_service.save_abbreviations([
            {"terme": "Simple", "abreviation": "SMP", "priorite": 2}
        ])
        
        exporter.create_export(rows, output_file_opt, optimize_designations=True)
        
        wb_opt = load_workbook(output_file_opt)
        ws_opt = wb_opt.active
        
        # Original designation must remain unchanged in the designation column (col 13)
        assert ws_opt.cell(row=3, column=13).value == "Simple Part with a very long designation name that exceeds 32 characters"
        # designation_erp (col 14) must contain the optimized value (uppercase, sans accent, abbreviated and truncated)
        assert ws_opt.cell(row=3, column=14).value == "SMP PART WITH A VERY LONG"
        
        # Let's also check row 2 (short designation "Top Assy" under 32 chars)
        # Original designation remains unchanged
        assert ws_opt.cell(row=2, column=13).value == "Top Assy"
        # designation_erp contains uppercase sans accents version
        assert ws_opt.cell(row=2, column=14).value == "TOP ASSY"
    finally:
        # Always restore main abbreviations list
        abbreviation_service.save_abbreviations(original_abbrevs)
    
    # Verify some styling (limited check)
    assert ws.cell(row=1, column=1).font.bold is True
    # Column 1 (Green) vs Column 13 (Orange)
    assert ws.cell(row=1, column=1).fill.start_color.index == "00CCFFCC"
    assert ws.cell(row=1, column=13).fill.start_color.index == "00FFC000"


def test_dynamic_columns_export(tmp_path):
    from pathlib import Path
    import json
    
    config_path = Path(__file__).parent.parent.parent / "ui" / "resources" / "columns_config.json"
    
    # Back up original configuration
    original_config = None
    if config_path.exists():
        try:
            with open(config_path, "r", encoding="utf-8") as f:
                original_config = json.load(f)
        except Exception:
            pass
            
    # Mock dynamic columns configuration
    mock_config = [
        {"header": "Mon_CAD", "source_type": "special_processed", "source_name": "special_cad", "default_value": "", "style": "plm"},
        {"header": "Couleur", "source_type": "solid_edge_property", "source_name": "couleur", "default_value": "Inconnue", "style": "cad"},
        {"header": "Mon_Niveau", "source_type": "special_processed", "source_name": "level", "default_value": "0", "style": "plm"}
    ]
    
    # Save mock config
    config_path.parent.mkdir(parents=True, exist_ok=True)
    with open(config_path, "w", encoding="utf-8") as f:
        json.dump(mock_config, f)
        
    try:
        output_file = str(tmp_path / "test_dynamic_export.xlsx")
        
        rows = [
            ExportRow(
                level=0, relationship="", order=1, quantity=1, repere="",
                special_cad="Root", plm_class="SUB_ASSY_A", ref_utilisat="Root",
                version="A", indice_1="-", indice_2="-", revision="1",
                designation="Top Assy", cus_createur="Admin", cus_date_crea="2026",
                user_version_1="", date_version_1="", matiere="",
                densite="", dia_se="", mode_appro="Fabrication interne", attachments="C:/Root.asm",
                custom_properties={}
            ),
            ExportRow(
                level=1, relationship="ComposedOf", order=2, quantity=2, repere="",
                special_cad="Part1", plm_class="PART_A", ref_utilisat="Part1",
                version="B", indice_1="A", indice_2="-", revision="1",
                designation="Simple Part", cus_createur="User", cus_date_crea="2026",
                user_version_1="", date_version_1="", matiere="Steel",
                densite="7.8", dia_se="", mode_appro="Fabrication interne", attachments="C:/Part1.par",
                custom_properties={"couleur": "Bleu"}
            )
        ]
        
        exporter = ExcelExporter()
        exporter.create_export(rows, output_file)
        
        assert os.path.exists(output_file)
        
        wb = load_workbook(output_file)
        ws = wb.active
        
        # Verify Headers
        assert ws.cell(row=1, column=1).value == "Mon_CAD"
        assert ws.cell(row=1, column=2).value == "Couleur"
        assert ws.cell(row=1, column=3).value == "Mon_Niveau"
        
        # Verify Data
        assert ws.cell(row=2, column=1).value == "Root"
        assert ws.cell(row=2, column=2).value == "Inconnue"  # default value
        assert ws.cell(row=2, column=3).value == 0
        
        assert ws.cell(row=3, column=1).value == "Part1"
        assert ws.cell(row=3, column=2).value == "Bleu"      # extracted custom value
        assert ws.cell(row=3, column=3).value == 1
        
        # Verify dynamic header colors: Column 1 is PLM (Orange), Column 2 is CAD (Green)
        assert ws.cell(row=1, column=1).fill.start_color.index == "00FFC000"  # Orange
        assert ws.cell(row=1, column=2).fill.start_color.index == "00CCFFCC"  # Green
        
    finally:
        # Restore original config
        if original_config is not None:
            with open(config_path, "w", encoding="utf-8") as f:
                json.dump(original_config, f, indent=2, ensure_ascii=False)
        elif config_path.exists():
            os.remove(config_path)
