from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from pathlib import Path
import argparse
import re
import unicodedata

# ==========================================================
# PARAMETRES
# ==========================================================

# Le fichier Excel a traiter n'est plus code en dur.
# Il est fourni au lancement du script, par exemple :
# python designation_plm_32_priorite_abreviations_COMPLET_V2.py "mon_fichier.xlsx"

NOM_ONGLET = "Structure"
COL_DESIGNATION = 13
LIMITE_CARACTERES = 32

def lire_arguments():
    parser = argparse.ArgumentParser(
        description="Corrige les designations PLM d'un fichier Excel en appliquant abreviations et tronquage a 32 caracteres."
    )
    parser.add_argument(
        "fichier_source",
        help="Chemin du fichier Excel a traiter (.xlsx)."
    )
    parser.add_argument(
        "-o",
        "--sortie",
        default=None,
        help="Chemin du fichier Excel de sortie. Par defaut : <nom_fichier>_CORRIGE.xlsx"
    )
    return parser.parse_args()

# Mots commerciaux / techniques a conserver en priorite lors du tronquage.
MOTS_A_PRESERVER = [
    "VINOLOK",
    "EMF",
    "GALAXY",
    "NEO",
]

# ==========================================================
# ABREVIATIONS
# ==========================================================
#
# Regles generales du dictionnaire PLM COSTRAL
# ----------------------------------------------------------
#
# 1. Une seule abreviation par mot.
#    Deux mots differents ne doivent pas partager la meme
#    abreviation afin d'eviter toute ambiguite dans le PLM,
#    les nomenclatures, les exports ERP et les recherches.
#
#    Exemples :
#       Bielle        -> BIEL
#       Biellette     -> BIELT
#       Module        -> MOD
#       Modelisation  -> MODL
#
# 2. Differencier les familles de mots.
#    Une machine, une fonction et une piece ne doivent pas
#    forcement utiliser la meme racine abregee.
#
#    Exemples :
#       Boucheuse  -> BOUC    (machine)
#       Bouchage   -> BOUCH   (fonction)
#       Bouchon    -> BCH     (piece)
#
#       Rinceuse   -> RINC    (machine)
#       Rincage    -> RINCG   (fonction)
#
#       Guide      -> GUID    (piece)
#       Guidage    -> GUIDG   (fonction)
#
# 3. Les pieces mecaniques utilisent de preference des
#    abreviations courtes de 3 a 5 caracteres, mais la clarte
#    reste prioritaire sur le gain d'un ou deux caracteres.
#
#    Exemples :
#       Longeron      -> LONG
#       Longitudinal  -> LONGI
#       Ressort       -> RES
#       Resine        -> RESI
#       Rotor         -> RTR
#       Rotule        -> ROT
#
# 4. Les noms commerciaux, gammes, marques et references
#    techniques importantes sont conserves autant que possible.
#    Ils peuvent aussi etre ajoutes dans MOTS_A_PRESERVER.
#
#    Exemples :
#       GALAXY
#       NEO
#       COMET
#       VINOLOK
#       ARDEASEAL
#       GUALA SEAL
#       HIGH TOP
#
# 5. Les designations sont abregees avant tout tronquage.
#    Le tronquage a 32 caracteres ne doit intervenir qu'apres
#    application de toutes les abreviations utiles.
#
# Chaque abbreviation est definie avec :
# (texte long, texte court, priorite)
#
# Priorite 1 = abbreviation importante, appliquee en premier.
# Priorite 2 = appliquee ensuite si la designation est encore trop longue.
# Priorite 3 = appliquee en dernier recours avant tronquage.
#
# A priorite identique, le script applique d'abord les abreviations
# qui font gagner le plus de caracteres.

# Les entrees ci-dessous sont issues de la mise a jour du dictionnaire
# fournie par le Bureau d'Etudes. Les variantes accentuees/non accentuees,
# les pluriels explicites et les termes separes par "/" sont generes
# automatiquement par la fonction construire_abreviations().

abreviations_source = [
    ('Afficheur', 'AFF', 2),
    ('Aimant', 'AIM', 2),
    ('Alimentation', 'ALIM', 2),
    ('Aluminium', 'ALU', 2),
    ('Amovible', 'AMOV', 2),
    ('Annulaire', 'ANN', 2),
    ('API', 'API', 2),
    ('Arbre', 'ARB', 2),
    ('Armoire', 'ARM', 2),
    ("Arrêt d'urgence", 'AU', 2),
    ('Arrière', 'AR', 2),
    ('Assemblé', 'ASM', 2),
    ('Automate', 'AUTO', 2),
    ('Avant', 'AV', 2),
    ('Axe', 'AX', 2),
    ('Bac', 'BAC', 2),
    ('Bague', 'BG', 2),
    ('Bandeau(x)', 'BAND', 3),
    ('Basculement', 'BASC', 1),
    ('Bielle', 'BIEL', 2),
    ('Biellette', 'BIELT', 2),
    ('Bille', 'BIL', 2),
    ('Bloc', 'BLK', 2),
    ('Bobine', 'BOB', 2),
    ('Boîtier', 'BOIT', 2),
    ('Borne', 'BOR', 2),
    ('Bornier', 'BORN', 2),
    ('Bouchage', 'BOUCH', 1),
    ('Boucheuse(s)', 'BOUC', 1),
    ('Bouchon', 'BCH', 1),
    ('Bout plat', 'BP', 2),
    ('Bouton', 'BTN', 2),
    ('Bras', 'BR', 2),
    ('Brushless', 'BLDC', 2),
    ('Butée', 'BUT', 2),
    ('Cable', 'CAB', 2),
    ('Câble', 'CAB', 2),
    ('Cadre', 'CAD', 2),
    ('Came', 'CAM', 2),
    ('Capot', 'CAP', 2),
    ('Capsuleuse', 'CAPS', 2),
    ('Capteur', 'CAPT', 2),
    ('Carte', 'CRT', 2),
    ('Centrale(s)', 'CENT', 2),
    ('Charnière', 'CHARN', 2),
    ('Chassis', 'CH', 1),
    ('Châssis', 'CH', 1),
    ('Circlips', 'CIR', 2),
    ('Circuit imprimé', 'CII', 2),
    ('Clavette', 'CLAV', 2),
    ('Codeur', 'COD', 2),
    ('Coffret', 'COFF', 2),
    ('Colonne', 'COL', 1),
    ('Commande', 'CMD', 2),
    ('Complet', 'CPL', 2),
    ('Compresseur', 'COMP', 2),
    ('Cône', 'CNE', 2),
    ('Connecteur', 'CONN', 2),
    ('Contacteur', 'CONT', 2),
    ('Convertisseur', 'CONV', 2),
    ('Cornière', 'CORN', 2),
    ('Corps', 'CPS', 2),
    ('Couronne', 'CRN', 2),
    ('Courroie', 'COUR', 2),
    ('Couvercle', 'COUV', 2),
    ('Couvercle(s)', 'COUV', 3),
    ('Crantée', 'CRANT', 2),
    ('Décalage', 'DEC', 2),
    ('Déclencheur', 'DECL', 2),
    ('Descente(s)', 'DESC', 2),
    ('Dessous', 'DSO', 2),
    ('Dessus', 'DSU', 2),
    ('Détecteur', 'DET', 2),
    ('Diamètre', 'Ø', 2),
    ('Différentiel', 'DIFF', 2),
    ('Diffuseur', 'DIFFU', 2),
    ('Disjoncteur', 'DISJ', 2),
    ('Disque', 'DISQ', 2),
    ('Dissipateur', 'DISS', 2),
    ('Distribution', 'DIST', 1),
    ('Doigt', 'DGT', 2),
    ('Douille', 'DOU', 2),
    ('Droit', 'DROI', 2),
    ('Écoulement', 'ECOU', 1),
    ('Égalisation', 'EGAL', 2),
    ('Électrique', 'ELEC', 2),
    ('Électrovanne', 'EV', 2),
    ('Embase', 'EMB', 2),
    ('Enfoncement(s)', 'ENF', 2),
    ('Ensemble', 'ENS', 1),
    ('Entraînement', 'ENTRN', 2),
    ('Entrée', 'ENT', 2),
    ('Entretoise', 'ENTR', 2),
    ('Équerre', 'EQR', 2),
    ('Étiquette', 'ETIQ', 2),
    ('Étoile', 'ETO', 2),
    ('Extension', 'EXTN', 2),
    ('Extérieur / Extérieure', 'EXT', 2),
    ('Faisceau', 'FAIS', 2),
    ('Fermeture', 'FERM', 2),
    ('Filtre', 'FILT', 2),
    ('Fixation', 'FIX', 1),
    ('Fonction', 'FCT', 2),
    ('Fond', 'FOND', 2),
    ('Fourreau', 'FOUR', 2),
    ('Galet', 'GAL', 2),
    ('Garniture', 'GARN', 2),
    ('Gauche', 'GAU', 2),
    ('Glissement(s)', 'GLIS', 2),
    ('Goulotte', 'GOUL', 2),
    ('Goupille', 'GOUP', 2),
    ('Goupillé', 'GOUP', 2),
    ('Gravée', 'GRAV', 2),
    ('Grille', 'GRIL', 2),
    ('Groupe', 'GRP', 2),
    ('Guidage', 'GUIDG', 2),
    ('Guide', 'GUID', 2),
    ('Habillage(s)', 'HABIL', 3),
    ('Hall', 'HALL', 2),
    ('Hauteur', 'HT', 2),
    ('High Top', 'HTP', 1),
    ('Horizontal', 'HOR', 2),
    ('Inductif', 'INDUC', 2),
    ('Inférieur / Inférieure', 'INF', 2),
    ('Injection', 'INJ', 2),
    ('Insert', 'INS', 2),
    ('Inter-châssis', 'ICH', 2),
    ('Interface', 'IF', 2),
    ('Intérieur / intérieure', 'INT', 2),
    ('Intermédiaire', 'INTM', 2),
    ('Interrupteur', 'INTERR', 2),
    ('Jaune/Noir', 'J/N', 2),
    ('Joint', 'JNT', 2),
    ('Joints', 'JNT', 2),
    ('Jonction', 'JONC', 2),
    ('Latéral', 'LAT', 2),
    ('Levier', 'LEV', 2),
    ('Longeron', 'LONG', 2),
    ('Longitudinal', 'LONGI', 2),
    ('Longue', 'LG', 2),
    ('Magasin', 'MAG', 2),
    ('Manomètre', 'MANO', 2),
    ('Modelisation', 'MODL', 2),
    ('Module', 'MOD', 2),
    ('Monté', 'MONT', 2),
    ('Moteur', 'MOT', 2),
    ('Motoréducteur', 'MR', 1),
    ('Niveau', 'NIV', 2),
    ('Noir', 'NR', 2),
    ('Oméga', 'OMG', 2),
    ('Option(s)', 'OPT', 3),
    ('Palier', 'PAL', 2),
    ('Panneau', 'PAN', 2),
    ('Panneau(x)', 'PAN', 2),
    ('Passe-fil', 'PF', 2),
    ('Patte', 'PAT', 2),
    ('Peigne', 'PEI', 2),
    ('Pied', 'PD', 2),
    ('Piste', 'PST', 2),
    ('Pivotement', 'PIV', 1),
    ('Plaque', 'PLQ', 2),
    ('Plaquette', 'PLAQ', 2),
    ('Plat', 'PLAT', 2),
    ('Platine', 'PLAT', 2),
    ('Plexiglas / Plexi', 'PLX', 2),
    ('Plot', 'PLT', 2),
    ('Portière', 'PORT', 2),
    ('Poteau', 'POT', 2),
    ('Potentiomètre', 'POTAR', 2),
    ('Poulie', 'POUL', 2),
    ('Poussoir', 'POUS', 2),
    ('Prémonté', 'PMT', 2),
    ('Prise', 'PRIS', 2),
    ('Profil', 'PROF', 2),
    ('Protection', 'PROT', 2),
    ('Pupitre', 'PUP', 2),
    ('Raccord', 'RAC', 2),
    ('Rail', 'RAIL', 2),
    ('Rallonge', 'RAL', 2),
    ('Rampe', 'RAMP', 2),
    ('Réception', 'RECP', 2),
    ('Reducteur', 'RED', 1),
    ('Réglage', 'REGL', 2),
    ('Renfort', 'RENF', 2),
    ('Renvoi', 'RENV', 2),
    ('Répartiteur', 'REP', 2),
    ('Repérage', 'REPER', 2),
    ('Résine', 'RESI', 2),
    ('Ressort', 'RES', 2),
    ('Rinçage', 'RINCG', 2),
    ('Rinceuse', 'RINC', 1),
    ('Rondelle(s)', 'RDL', 2),
    ('Rotatif', 'ROTAT', 2),
    ('Rotor', 'RTR', 2),
    ('Rotule', 'ROT', 2),
    ('Roulement', 'RLT', 2),
    ('Roulette', 'ROUL', 2),
    ('Sans tête', 'ST', 2),
    ('Sécurité', 'SEC', 2),
    ('Semelle', 'SEM', 2),
    ('Séparation', 'SEP', 2),
    ('Servo-moteur', 'SM', 2),
    ('Socle', 'SOC', 2),
    ('Sortie', 'SORT', 2),
    ('Stabilisée', 'STAB', 2),
    ('Standard', 'STD', 2),
    ('Structure', 'STR', 2),
    ('Supérieure(s)', 'SUP', 2),
    ('Support(s)', 'SUPP', 2),
    ('Surmoulé', 'SURM', 2),
    ('Symétrique', 'SYM', 2),
    ('Tendeur', 'TEND', 2),
    ('TETE BOIS / TÊTE BOIS', 'TB', 1),
    ('Tireuse', 'TIR', 1),
    ('Tôle', 'TOL', 2),
    ('Transformateur', 'TRANSFO', 2),
    ('Transporteur', 'TRANSP', 2),
    ('Traverse(s)', 'TRAV', 2),
    ('Tube', 'TUB', 2),
    ('Variateur', 'VAR', 2),
    ('Ventilateur', 'VENT', 2),
    ('Verrou', 'VER', 2),
    ('Vertical', 'VERT', 2),
    ('Verticale', 'VERT', 2),
    ('Voyant', 'VOY', 2),
]


def _supprimer_accents_local(texte):
    texte = unicodedata.normalize("NFD", str(texte))
    return "".join(
        caractere
        for caractere in texte
        if unicodedata.category(caractere) != "Mn"
    )


def _variantes_terme(terme):
    """Genere les variantes utiles : accents, pluriels explicites, formes separees par /."""
    terme = nettoyer_espaces(terme) if "nettoyer_espaces" in globals() else re.sub(r"\s+", " ", str(terme)).strip()
    variantes = []

    morceaux = [m.strip() for m in terme.split("/") if m.strip()] if " / " in terme else [terme]

    for morceau in morceaux:
        candidats = [morceau]

        # Gestion des notations de type Bandeau(x), Support(s), Rondelle(s), etc.
        if "(x)" in morceau:
            candidats.append(morceau.replace("(x)", ""))
            candidats.append(morceau.replace("(x)", "x"))
        if "(s)" in morceau:
            candidats.append(morceau.replace("(s)", ""))
            candidats.append(morceau.replace("(s)", "s"))

        for candidat in candidats:
            candidat = candidat.replace("(x)", "").replace("(s)", "")
            candidat = re.sub(r"\s+", " ", candidat).strip()
            if not candidat:
                continue
            variantes.append(candidat)
            variantes.append(_supprimer_accents_local(candidat))

    # Dedoublonnage en conservant l'ordre
    resultat = []
    vus = set()
    for variante in variantes:
        cle = variante.casefold()
        if cle not in vus:
            vus.add(cle)
            resultat.append(variante)
    return resultat


def construire_abreviations(source):
    abreviations = []
    vus = set()
    for terme, abreviation, priorite in source:
        for variante in _variantes_terme(terme):
            cle = (variante.casefold(), str(abreviation).casefold(), priorite)
            if cle in vus:
                continue
            vus.add(cle)
            abreviations.append((variante, abreviation, priorite))
    return abreviations


abreviations = construire_abreviations(abreviations_source)

# Precompilation des expressions regulieres.
# Tri : priorite croissante, puis gain de caracteres decroissant.
abreviations_compilees = []
for texte_long, texte_court, priorite in sorted(
    abreviations,
    key=lambda item: (item[2], -(len(item[0]) - len(item[1]))),
):
    abreviations_compilees.append(
        {
            "motif": re.compile(rf"\b{re.escape(texte_long)}\b", flags=re.IGNORECASE),
            "remplacement": texte_court,
            "texte_long": texte_long,
            "priorite": priorite,
            "gain": len(texte_long) - len(texte_court),
        }
    )

# ==========================================================
# FONCTIONS
# ==========================================================

def nettoyer_espaces(texte):
    return re.sub(r"\s+", " ", str(texte)).strip()


def supprimer_accents(texte):
    """
    Supprime les accents du texte.
    Exemple : Ecoulement arriere -> Ecoulement arriere
    """
    texte = unicodedata.normalize("NFD", str(texte))
    texte = "".join(
        caractere
        for caractere in texte
        if unicodedata.category(caractere) != "Mn"
    )
    return texte


def majuscules_sans_accents(texte):
    """
    Convertit le texte en MAJUSCULES sans accents.
    Exemple : Ecoulement arriere -> ECOULEMENT ARRIERE
    """
    return supprimer_accents(texte).upper()


def appliquer_abreviations_prioritaires(texte, limite=LIMITE_CARACTERES):
    """
    Applique les abreviations avant de tronquer.

    Regle de decision :
    1. On ne touche pas aux textes deja sous la limite.
    2. Tant que le texte est trop long, on cherche les abreviations possibles.
    3. On choisit d'abord la priorite la plus forte : 1 avant 2 avant 3.
    4. A priorite identique, on choisit celle qui fait gagner le plus de caracteres.
    5. On ne tronque qu'apres avoir epuise toutes les abreviations utiles.
    """
    texte = nettoyer_espaces(texte)

    if len(texte) <= limite:
        return texte, False

    texte_modifie = texte
    modifie = False

    while len(texte_modifie) > limite:
        meilleurs_choix = []

        for regle in abreviations_compilees:
            nouveau_texte = regle["motif"].sub(
                regle["remplacement"],
                texte_modifie,
                count=1,
            )

            if nouveau_texte == texte_modifie:
                continue

            nouveau_texte = nettoyer_espaces(nouveau_texte)
            gain_reel = len(texte_modifie) - len(nouveau_texte)

            if gain_reel <= 0:
                continue

            meilleurs_choix.append(
                (
                    regle["priorite"],
                    -gain_reel,
                    nouveau_texte,
                )
            )

        if not meilleurs_choix:
            break

        meilleurs_choix.sort()
        texte_modifie = meilleurs_choix[0][2]
        modifie = True

    return texte_modifie, modifie


def tronquer_mots_entiers(texte, limite=LIMITE_CARACTERES):
    """
    Tronque sans couper les mots.
    """
    texte = nettoyer_espaces(texte)

    if len(texte) <= limite:
        return texte

    mots = texte.split(" ")
    resultat = ""

    for mot in mots:
        candidat = mot if resultat == "" else f"{resultat} {mot}"

        if len(candidat) <= limite:
            resultat = candidat
        else:
            break

    return resultat


def trouver_mot_a_preserver(texte):
    """
    Retourne le premier mot-cle a preserver trouve dans le texte.
    """
    for mot in MOTS_A_PRESERVER:
        if re.search(rf"\b{re.escape(mot)}\b", texte, flags=re.IGNORECASE):
            return mot
    return None


def tronquer_en_conservant_mot_cle(texte, mot_cle, limite=LIMITE_CARACTERES):
    """
    Tronque sans couper les mots, en conservant obligatoirement le mot-cle.

    Exemple : si GALAXY est present, le resultat final conservera GALAXY.
    Le script garde autant de mots que possible avant le mot-cle.
    """
    texte = nettoyer_espaces(texte)

    match = re.search(
        rf"\b{re.escape(mot_cle)}\b",
        texte,
        flags=re.IGNORECASE,
    )

    if not match:
        return tronquer_mots_entiers(texte, limite)

    resultat = match.group(0)
    avant = texte[:match.start()].strip()
    mots_avant = avant.split() if avant else []

    for mot in reversed(mots_avant):
        candidat = f"{mot} {resultat}"

        if len(candidat) <= limite:
            resultat = candidat
        else:
            break

    return resultat


def optimiser_designation(texte):
    """
    1. Nettoie les espaces.
    2. Applique les abreviations selon leur priorite.
    3. A priorite identique, applique l'abbreviation avec le plus grand gain.
    4. Si le texte reste trop long, tronque sans couper les mots.
    5. Si un mot-cle protege est present, le conserve en priorite.
    """
    texte = nettoyer_espaces(texte)

    if len(texte) <= LIMITE_CARACTERES:
        return texte, False

    texte_abrege, abrege = appliquer_abreviations_prioritaires(texte)

    if len(texte_abrege) <= LIMITE_CARACTERES:
        return texte_abrege, True

    mot_cle = trouver_mot_a_preserver(texte_abrege)
    if mot_cle:
        return tronquer_en_conservant_mot_cle(texte_abrege, mot_cle), True

    return tronquer_mots_entiers(texte_abrege), True

# ==========================================================
# TRAITEMENT EXCEL
# ==========================================================

def traiter_fichier_excel(fichier_source, fichier_sortie=None):
    source = Path(fichier_source)

    if not source.exists():
        raise FileNotFoundError(f"Fichier introuvable : {source}")

    if fichier_sortie is None:
        fichier_sortie = source.parent / f"{source.stem}_CORRIGE{source.suffix}"
    else:
        fichier_sortie = Path(fichier_sortie)

    wb = load_workbook(source)

    if NOM_ONGLET not in wb.sheetnames:
        raise ValueError(f"Onglet introuvable : {NOM_ONGLET}")

    ws = wb[NOM_ONGLET]

    rouge = PatternFill(
        fill_type="solid",
        fgColor="FF9999",
    )

    nb_modifications = 0

    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=COL_DESIGNATION)

        if not isinstance(cell.value, str):
            continue

        ancienne_valeur = cell.value
        nouvelle_valeur, modifie = optimiser_designation(ancienne_valeur)
        valeur_finale = majuscules_sans_accents(nouvelle_valeur)

        if valeur_finale != ancienne_valeur:
            cell.value = valeur_finale
            nb_modifications += 1

            # On ne colorie en rouge que si la designation a reellement subi une abreviation/tronquage
            if modifie and nouvelle_valeur != ancienne_valeur:
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=row, column=col).fill = rouge

    wb.save(fichier_sortie)

    print(f"Lignes modifiees : {nb_modifications}")
    print(f"Fichier genere : {fichier_sortie}")

    return fichier_sortie, nb_modifications


if __name__ == "__main__":
    args = lire_arguments()
    traiter_fichier_excel(args.fichier_source, args.sortie)
